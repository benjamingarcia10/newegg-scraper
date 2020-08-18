import requests
import openpyxl
from bs4 import BeautifulSoup as soup
import re
from captcha import solve_captcha

# Get search type for scraping (by query or link)
search_type = input('Would you like to search by query (q) or link (l)? ').lower().strip()
while not (search_type == 'q' or search_type == 'l'):
    print('Invalid input.')
    search_type = input('Would you like to search by query (q) or link (l)? ').lower().strip()
if search_type == 'q':
    query = input('What item would you like to search for on Newegg.com? ').lower().strip()
    search_url = f'https://www.newegg.com/p/pl?d={query}'
elif search_type == 'l':
    link = input('Please insert a Newegg.com initial link to scrape: ').lower().strip()
    while not 'newegg.com' in link:
        print('Invalid Newegg link.')
        link = input('Please insert a Newegg.com link to scrape: ').lower().strip()
    search_url = link

# Get response to save data to spreadsheet or not
save_data = input('Would you like to save scraped data? (y or n)? ').lower().strip()
while not (save_data == 'y' or save_data == 'n'):
    print('Invalid input.')
    save_data = input('Would you like to save scraped data? (y or n)? ').lower().strip()
if save_data == 'y':
    is_data_saved = True
elif save_data == 'n':
    is_data_saved = False

# String to set data to if info is not found on the webpage
unfound_data_string = 'Not found'

# Get response from initial Newegg URL
page_headers = {
    'authority': 'www.newegg.com',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.105 Safari/537.36',
    'method': 'GET',
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'accept-encoding': 'gzip, deflate, br',
    'accept-language': 'en-US,en;q=0.9'
}
initial_response = requests.get(search_url, headers=page_headers)
response_html = initial_response.text

# Parse HTML
initial_soup = soup(response_html, 'html.parser')

# Check for CAPTCHA
if 'Are you a human?' in initial_soup.text:
    print('CAPTCHA REQUIRED! Launching browser.')
    initial_soup = solve_captcha(initial_soup, initial_response.url)

# Get number of pages for results
pages = initial_soup.find(class_='list-tool-pagination-text')
page_amount = 1
if pages is not None:
    page_amount = int(pages.text.split("/", 1)[1])
    print(f'Page Count: {page_amount}')

# Ask how many pages to scrape for
max_page_amount = 1
if page_amount > 1:
    max_page_amount = input('What page would you like to scrape until (inclusive)? ').strip()
    while not isinstance(max_page_amount, int):
        try:
            max_page_amount = int(max_page_amount)
        except:
            print('Invalid page amount.')
            max_page_amount = input('What page would you like to scrape until (inclusive)? ').strip()
    if max_page_amount > page_amount:
        max_page_amount = page_amount
    elif max_page_amount < 1:
        max_page_amount = 1
print(f'Scraping {max_page_amount} pages.\n')

# All features found via webscraping (some products may not have certain features)
list_of_all_features = []
# All items found across all pages
items = []

# Check items on all pages
# for page_number in range(1, page_amount + 1):
for page_number in range(1, max_page_amount + 1):
    amount_of_items_on_page = 0

    if 'ID-' in initial_response.url and page_number > 1:
        split_url = re.compile('(ID-[0-9]+)').split(initial_response.url)
        page_url = f'{split_url[0]}{split_url[1]}/Page-{page_number}'
    elif 'ID-' in initial_response.url and page_number == 1:
        page_url = f'{initial_response.url}'
    else:
        page_url = f'{initial_response.url}&page={page_number}'
        # page_url = f'{initial_response.url}&PageSize=96&page={page_number}'
    print(f'Scraping Page {page_number}/{page_amount}: {page_url}')

    # Get response from desired Newegg page URL
    page_response = requests.get(page_url, headers=page_headers)
    page_html = page_response.text

    # Parse HTML
    page_soup = soup(page_html, 'html.parser')

    # Check for CAPTCHA
    if 'Are you a human?' in page_soup.text:
        print('\tCAPTCHA REQUIRED! Launching browser.')
        page_soup = solve_captcha(page_soup, page_response.url)

    # Get item containers
    containers = page_soup.find_all('div', class_='item-container')

    # Check all items on the page and retrieve item data stored in Python dictionary
    # All data in item data is key:single value except for the key "feature_list"
    #       which maps to another Python dictionary (feature_name:feature_value)
    for container in containers:
        item_data = {}

        # Get brand of item from image link
        try:
            brand = container.find(class_='item-brand').img['title']
        except:
            brand = unfound_data_string
        finally:
            item_data['brand'] = brand

        # Get name and page url of item
        try:
            name = container.find(class_='item-title').text.replace('\n', ' | ').strip()
            item_url = container.find(class_='item-title')['href']
        except:
            name = unfound_data_string
            item_url = unfound_data_string
        finally:
            item_data['name'] = name
            item_data['url'] = item_url

        # Get list containing all item features
        try:
            feature_containers = container.find(class_='item-features').find_all('li')
            feature_list = {}
            for feature in feature_containers:
                feature_split = feature.text.replace('\n', ' | ').split(':', 1)
                feature_list[feature_split[0].strip()] = feature_split[1].strip()
                if feature_split[0].strip() not in list_of_all_features:
                    list_of_all_features.append(feature_split[0].strip())
        except:
            feature_list = {}
        finally:
            item_data['feature_list'] = feature_list

        # Check for item promotion
        try:
            item_promotion = container.find(class_='item-promo').text.replace('\n', ' | ').strip()
        except:
            item_promotion = unfound_data_string
        finally:
            item_data['item_promotion'] = item_promotion

        # Try to get item cost
        try:
            item_price_tag = container.find(class_='price-current')
            item_price_dollars = item_price_tag.strong.text.replace('\n', ' | ').strip()
            item_price_cents = item_price_tag.sup.text.replace('\n', ' | ').strip()
            item_price = f'${item_price_dollars}{item_price_cents}'
        except:
            item_price = unfound_data_string
        finally:
            item_data['item_price'] = item_price

        # Check if item is shipped by newegg
        if container.find(class_='shipped-by-newegg') is not None:
            is_shipped_by_newegg = True
        else:
            is_shipped_by_newegg = False
        item_data['is_shipped_by_newegg'] = is_shipped_by_newegg

        # Get shipping cost of item
        try:
            shipping_cost = container.find(class_='price-ship').text.replace('\n', ' | ').strip()
        except:
            shipping_cost = unfound_data_string
        finally:
            item_data['shipping_cost'] = shipping_cost

        items.append(item_data)
        amount_of_items_on_page += 1
    print(f'\tNumber of Items Found: {amount_of_items_on_page}\n')

if is_data_saved:
    # Save all items to items.xlsx
    # Item Data:
    # items[x]['brand'] = brand of item
    # items[x]['name'] = name of item
    # items[x]['url'] = url to the item page
    # items[x]['feature_list'] = feature list of item stored in Python dictionary
    # items[x]['item_promotion'] = applicable item promotion
    # items[x]['item_price'] = cost of item
    # items[x]['is_shipped_by_newegg'] = True/False if item is shipped by Newegg
    # items[x]['shipping_cost'] = cost of shipping

    wb = openpyxl.Workbook()
    ws = wb.active
    # worksheet_headers = ['Brand', 'Name', 'URL', list_of_all_features, 'Promotion',
    #                      'Price', 'Shipped By Newegg?', 'Shipping Cost']
    worksheet_headers = {
        'Brand': 'brand',
        'Name': 'name',
        'URL': 'url',
        'Features': 'feature_list',
        'Promotion': 'item_promotion',
        'Price': 'item_price',
        'Shipped By Newegg?': 'is_shipped_by_newegg',
        'Shipping Cost': 'shipping_cost'
    }

    mapped_worksheet_headers = {}
    header_row = 1
    column_count = 1
    for header in worksheet_headers:
        if header == 'Features':
            for feature in list_of_all_features:
                mapped_worksheet_headers[feature] = {
                    'column': column_count,
                    'tag': worksheet_headers.get(header),
                    'feature': feature
                }
                ws.cell(row=header_row, column=column_count).value = feature
                column_count += 1
        else:
            mapped_worksheet_headers[header] = {
                'column': column_count,
                'tag': worksheet_headers.get(header)
            }
            ws.cell(row=header_row, column=column_count).value = header
            column_count += 1

    item_row = 2
    for item in items:

        for header in mapped_worksheet_headers.values():
            try:
                ws.cell(row=item_row, column=header['column']).value = item[header['tag']]
            except:
                try:
                    ws.cell(row=item_row, column=header['column']).value = item[header['tag']].get(header['feature'])
                except:
                    ws.cell(row=item_row, column=header['column']).value = unfound_data_string

        try:
            url_cell = ws.cell(row=item_row, column=mapped_worksheet_headers.get('URL')['column'])
            url_cell.hyperlink = item[mapped_worksheet_headers.get('URL')['tag']]
            url_cell.style = 'Hyperlink'
        except:
            pass
        item_row += 1

    wb.save(f'items.xlsx')
    print('Data saved to items.xlsx')
else:
    for item in items:
        print(item)
    print(f'{len(items)} items found.')